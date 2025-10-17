#!/usr/bin/env python3
"""Import a 30-question quiz bank from a Microsoft Forms export."""
from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sqlite3
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence

try:  # Optional dependency
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - openpyxl may be unavailable
    load_workbook = None

try:
    from zoneinfo import ZoneInfo

    TZ = ZoneInfo("Asia/Kuala_Lumpur")

    def now_str() -> str:
        return datetime.now(TZ).strftime("%Y-%m-%d_%H%M")

    def timestamp_str() -> str:
        return datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
except Exception:  # pragma: no cover - zoneinfo not available on some systems
    TZ = None

    def now_str() -> str:
        return datetime.now().strftime("%Y-%m-%d_%H%M")

    def timestamp_str() -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
BACKUP_DIR = ROOT / "backups"
DEFAULT_DB = ROOT / "pla.db"

TOPIC_FUNDAMENTALS = "Data Modeling & DBMS Fundamentals"
TOPIC_NORMALIZATION = "Normalization & Dependencies"

NORMALIZATION_KEYWORDS = [
    "normalization",
    "1nf",
    "2nf",
    "3nf",
    "functional dependency",
    "fd",
    "partial dependency",
    "transitive dependency",
    "bcnf",
    "anomaly",
    "decomposition",
    "lossless",
    "determinant",
    "prime attribute",
    "closure",
    "cover",
    "dependency",
    "normal form",
]

FUNDAMENTALS_KEYWORDS = [
    "dbms",
    "database",
    "table",
    "row",
    "column",
    "primary key",
    "foreign key",
    "candidate key",
    "entity",
    "relationship",
    "attribute",
    "cardinality",
    "schema",
    "er",
    "diagram",
    "sql",
    "ddl",
    "dml",
    "data model",
    "relation",
    "tuple",
    "domain",
]

POINTS_PREFIX = "Points - "

QuestionRecord = Dict[str, object]


def _trim(text: object) -> str:
    return str(text).strip() if text is not None else ""


class ImportErrorWithReport(RuntimeError):
    """Raised when the incoming data cannot be converted to 30 questions."""


def parse_arguments(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "path",
        nargs="?",
        help="Path to Microsoft Forms export (.xlsx preferred, .csv fallback)",
    )
    return parser.parse_args(argv)


def detect_input_file(path_argument: Optional[str]) -> Path:
    if path_argument:
        candidate = Path(path_argument).expanduser().resolve()
        if not candidate.exists():
            raise SystemExit(f"Input file not found: {candidate}")
        return candidate

    if load_workbook is not None:
        xlsx_files = sorted(DATA_DIR.glob("*.xlsx"))
        if xlsx_files:
            return xlsx_files[0]
    else:
        xlsx_files = sorted(DATA_DIR.glob("*.xlsx"))
        if xlsx_files:
            print(
                "openpyxl is not installed; ignoring XLSX files and falling back to CSV.",
                file=sys.stderr,
            )
    csv_files = sorted(DATA_DIR.glob("*.csv"))
    if csv_files:
        return csv_files[0]

    raise SystemExit("No input file supplied and none found in app/data.")


def read_rows_from_xlsx(path: Path) -> List[List[str]]:
    if load_workbook is None:
        raise SystemExit("openpyxl is required to read XLSX files.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet = workbook.active
    rows: List[List[str]] = []
    for excel_row in sheet.iter_rows(values_only=True):
        values = [_trim(cell) for cell in excel_row]
        if any(values):
            rows.append(values)
    workbook.close()
    return rows


def read_rows_from_csv(path: Path) -> List[List[str]]:
    rows: List[List[str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for csv_row in reader:
            values = [_trim(cell) for cell in csv_row]
            if any(values):
                rows.append(values)
    return rows


def read_input_table(path: Path) -> List[List[str]]:
    if path.suffix.lower() == ".xlsx":
        return read_rows_from_xlsx(path)
    if path.suffix.lower() == ".csv":
        return read_rows_from_csv(path)
    raise SystemExit(f"Unsupported input format: {path.suffix}")


def build_question_map(table: List[List[str]]) -> Dict[str, Dict[str, object]]:
    if not table:
        raise ImportErrorWithReport("Input file contains no data rows.")

    header = [_trim(cell) for cell in table[0]]
    if not any(header):
        raise ImportErrorWithReport("Header row is empty.")

    question_columns: "OrderedDict[str, Dict[str, int]]" = OrderedDict()
    for idx, column_name in enumerate(header):
        if not column_name or not column_name.startswith(POINTS_PREFIX):
            continue

        question_text = column_name[len(POINTS_PREFIX) :].strip()
        if not question_text or question_text in question_columns:
            continue

        try:
            answer_index = header.index(question_text)
        except ValueError:
            continue

        question_columns[question_text] = {
            "answer_idx": answer_index,
            "points_idx": idx,
        }

    if not question_columns:
        raise ImportErrorWithReport("Could not locate any 'Points - <Question>' columns.")

    details: "OrderedDict[str, Dict[str, object]]" = OrderedDict()
    for question_text, indices in question_columns.items():
        details[question_text] = {
            "question": question_text,
            "answer_idx": indices["answer_idx"],
            "points_idx": indices["points_idx"],
            "options": [],
            "correct_counts": {},
        }

    for raw_row in table[1:]:
        if not raw_row:
            continue
        row = list(raw_row)
        if len(row) < len(header):
            row.extend([""] * (len(header) - len(row)))
        for question_text, info in details.items():
            answer_idx = info["answer_idx"]
            points_idx = info["points_idx"]
            answer_val = row[answer_idx] if answer_idx < len(row) else ""
            points_val = row[points_idx] if points_idx < len(row) else ""

            answer_text = _trim(answer_val)
            if not answer_text:
                continue

            options: List[str] = info["options"]  # type: ignore[assignment]
            if answer_text not in options:
                if len(options) >= 6:
                    continue
                options.append(answer_text)

            counts: Dict[str, int] = info["correct_counts"]  # type: ignore[assignment]
            if parse_points(points_val):
                counts[answer_text] = counts.get(answer_text, 0) + 1

    return details


def parse_points(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return 1 if float(value) > 0 else 0
    text = _trim(value)
    if not text:
        return 0
    try:
        return 1 if float(text) > 0 else 0
    except ValueError:
        lowered = text.lower()
        if lowered in {"true", "yes"}:
            return 1
        return 0


def choose_correct_option(options: List[str], counts: Dict[str, int]) -> str:
    if not options:
        raise ImportErrorWithReport("A question has no answer options.")
    best_option = options[0]
    best_score = counts.get(best_option, 0)
    for option in options[1:]:
        score = counts.get(option, 0)
        if score > best_score:
            best_option = option
            best_score = score
    return best_option


def classify_topic(question: str) -> str:
    lowered = question.lower()
    if any(keyword in lowered for keyword in NORMALIZATION_KEYWORDS):
        return TOPIC_NORMALIZATION
    if any(keyword in lowered for keyword in FUNDAMENTALS_KEYWORDS):
        return TOPIC_FUNDAMENTALS
    return TOPIC_FUNDAMENTALS


def build_question_records(table: List[List[str]]) -> List[QuestionRecord]:
    details = build_question_map(table)
    questions: List[QuestionRecord] = []
    for question_text, info in details.items():
        options: List[str] = list(info["options"])  # type: ignore[assignment]
        if not options:
            raise ImportErrorWithReport(
                f"Question '{question_text[:60]}' has no discovered answer options."
            )
        if len(options) > 6:
            options = options[:6]

        # Microsoft Forms typically supplies four options, but occasionally a
        # dataset can contain fewer (e.g., true/false). We keep whatever was
        # observed and map letters in discovery order so callers can still work
        # with the reduced option set while staying within the 4–6 guidance.
        counts: Dict[str, int] = info["correct_counts"]  # type: ignore[assignment]
        correct_option = choose_correct_option(options, counts)
        if correct_option not in options:
            options.insert(0, correct_option)
        letters = {option: chr(ord("A") + idx) for idx, option in enumerate(options)}
        correct_letter = letters.get(correct_option, "A")
        topic = classify_topic(question_text)
        explanation = (
            f"Review the module for {topic}. Focus on key definitions and examples."
        )
        record: QuestionRecord = {
            "question": question_text,
            "options": options,
            "correct_letter": correct_letter,
            "two_category": topic,
            "explanation": explanation,
        }
        questions.append(record)
    return questions


def ensure_thirty_questions(records: List[QuestionRecord]) -> None:
    if len(records) != 30:
        lines = [
            f"Parsed {len(records)} questions; expected 30. Detailed mapping:",
        ]
        for record in records:
            short_question = record["question"].replace("\n", " ")[:70]
            options = record["options"]  # type: ignore[assignment]
            lines.append(
                f" - {short_question} ({len(options)} options, topic={record['two_category']})"
            )
        raise ImportErrorWithReport("\n".join(lines))


def backup_database(db_path: Path) -> None:
    if not db_path.exists():
        print(f"Database not found at {db_path}. Continuing without backup.")
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = now_str()
    destination = BACKUP_DIR / f"pla_{stamp}.db"
    shutil.copy2(db_path, destination)
    print(f"Backup created at {destination}")


def connect_database(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def import_questions(db_path: Path, records: List[QuestionRecord]) -> None:
    conn = connect_database(db_path)
    try:
        conn.execute("BEGIN")
        conn.execute("DELETE FROM response")
        conn.execute("DELETE FROM attempt")
        conn.execute("DELETE FROM quiz")
        insert_sql = (
            "INSERT INTO quiz (question, options_text, correct_answer, two_category, explanation) "
            "VALUES (?, ?, ?, ?, ?)"
        )
        for record in records:
            options = record["options"]  # type: ignore[assignment]
            conn.execute(
                insert_sql,
                (
                    record["question"],
                    json.dumps(options, ensure_ascii=False),
                    record["correct_letter"],
                    record["two_category"],
                    record["explanation"],
                ),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def summarize(records: List[QuestionRecord]) -> None:
    fundamentals = sum(1 for record in records if record["two_category"] == TOPIC_FUNDAMENTALS)
    normalization = sum(1 for record in records if record["two_category"] == TOPIC_NORMALIZATION)
    print("Import complete.")
    print(f"Total inserted: {len(records)}")
    print(f" - {TOPIC_FUNDAMENTALS}: {fundamentals}")
    print(f" - {TOPIC_NORMALIZATION}: {normalization}")


def main(argv: Optional[Sequence[str]] = None) -> None:
    args = parse_arguments(argv)
    input_path = detect_input_file(args.path)
    print(f"Reading: {input_path}")
    table = read_input_table(input_path)
    try:
        records = build_question_records(table)
        ensure_thirty_questions(records)
    except ImportErrorWithReport as exc:
        print(exc)
        return

    db_env = os.getenv("PLA_DB")
    db_path = Path(db_env).expanduser().resolve() if db_env else DEFAULT_DB
    print(f"Target database: {db_path}")
    backup_database(db_path)
    import_questions(db_path, records)
    summarize(records)


if __name__ == "__main__":
    main()
